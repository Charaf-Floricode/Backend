#!/usr/bin/env python3
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from sqlalchemy import create_engine, text, inspect
import csv
# --- CONFIGURATIE ---
TEMPLATE_PATH = Path(r"C:\Users\Floricode\Desktop\GPC code V2\GPC\Import_Florecompc2_template.xlsx")
DATA_DIR      = Path(r"C:\Users\Floricode\Desktop\GPC code V2\GPC\Data")
OUTPUT_DIR      = Path(r"C:\Users\Floricode\Desktop\GPC code V2\GPC\Output")
FILLED_PATH   = Path(r"C:\Users\Floricode\Desktop\GPC code V2\GPC\Import_Florecompc2_filled.xlsx")

DB_CONNECTION_STRING = (
    "postgresql+psycopg2://floricode_user:"
    "Hoofddorp123@localhost:5432/floricode"
)

# 1) Sheet → bestands-prefix
SHEETS_WITH_CODELIST = {
    "Benaming":        "CN??????.txt",
    "Benamingstype":   "CM??????.txt",
    "Cultivar":        "CC??????.txt",
    "Geslacht":        "CG??????.txt",
    "Gewas":           "CT??????.txt",
    "Groep":           "CO??????.txt",
    "Kenmerkgroep":    "CU??????.txt",
    "Kenmerktype":     "CE??????.txt",
    "Kenmerkwaarde":   "CV??????.txt",
    "Product":         "CP??????.txt",
    "Productkenmerk":  "CF??????.txt",
    "Regl. Kenmerktype":"CY??????.txt",
    "Soort":           "CS??????.txt",
    "Toepassing":      "CA??????.txt",
    "Voorschrift type":"CR??????.txt",
}

# 2) Sheet → Postgres-tabel
SHEET_TO_TABLE = {
    "Benaming":        "NAME_GPC",
    "Benamingstype":   "NAME_TYPE",
    "Cultivar":        "PLANT",
    "Geslacht":        "_import_GENUS",
    "Gewas":           "_import_PLANT",
    "Groep":           "GROUP",
    "Kenmerkgroep":    "SEGMENT_FAMILY_CLASS_BRICK",
    "Kenmerktype":     "ATTRIBUTE_TYPE",
    "Kenmerkwaarde":   "ATTRIBUTE_VALUE",
    "Product":         "PRODUCT_GPC",
    "Productkenmerk":  "BRICK_ATTRIBUTE_TYPE_ATTRIBUTE_VALUE",
    "Regl. Kenmerktype":"NAME_TYPE",
    "Soort":           "_import_SPECIES",
    "Toepassing":      "APPLICATION",
    "Voorschrift type":"CR_IMPORT"
}

# 3) Nederlandse → Engelse kolomnamen
DUTCH_TO_EN = {
    "codelijst":            "code_list_id",
    "registratie_autoriteit":"registrar_id",
    "gewas_id":             "plant_registration_number",
    "geslacht_id":          "genus_id",
    "soort_id":             "species_id",
    "cultivar_id":          "cultivar_id",
    "ingangsdatum":         "entry_date",
    "eindigingsdatum":      "expiry_date",
    "wijzigings_tijdstip":  "change_date_time",
    "latijnse_geslachtsnaam":"latin_genus_name",
    "naam": "latin_genus_name",
    "gecombineerd_product": "combined_product",
    "latijnse_soortnaam":    "latin_species_name"
    # … vul aan indien nodig …
}

START_ROW = 4
START_COL = 1
def normalize_and_translate(df: pd.DataFrame) -> pd.DataFrame:
    # 1) lowercase + underscores
    df.columns = (
        df.columns
          .str.strip()
          .str.lower()
          .str.replace(" ", "_")
          .str.replace(r"[^\w_]", "", regex=True)
    )
    # 2) Dutch → English
    return df.rename(columns=DUTCH_TO_EN)

# 4) View-definitie
VIEW_DDLS = {
    "view_plant_genus_species": """
    CREATE OR REPLACE VIEW "Plant_Genus_Species" AS
    SELECT
      p.plant_registration_number,
      p.genus_id,
      g.latin_genus_name,
      p.species_id,
      s.latin_species_name,
      CASE
        WHEN RIGHT(s.latin_species_name,3)='Grp' THEN g.latin_genus_name
        ELSE CONCAT(g.latin_genus_name,' ',s.latin_species_name)
      END AS genus_species
    FROM "_import_PLANT"    AS p
    JOIN "_import_GENUS"    AS g ON p.genus_id = g.genus_id
    LEFT JOIN "_import_SPECIES" AS s
      ON p.plant_registration_number = s.plant_registration_number
     AND p.species_id              = s.species_id;
    """
}

UPDATE_QUERIES = {
    "01_PGSP_GS_RegiD": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = b.brick_code
        FROM "_import_Bricks_With_Genus_Species" b
        JOIN "_import_Groups_With_Genus_Species" g
          ON b."gpc_application" = g."gpc_application"
        JOIN "Plant_Genus_Species" s
          ON b.genus_species = s.genus_species
        WHERE b."match_on_x" = 'GS'
          AND p.group_code = g.group_code
          AND p.plant_registration_number = s.plant_registration_number;
    """,
   "Q2_GS_null": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = b.brick_code
        FROM "Plant_Genus_Species"       AS s,
             "_import_Groups_With_Genus_Species" AS g,
             "_import_Bricks_With_Genus_Species" AS b
        WHERE p.plant_registration_number = s.plant_registration_number
          AND p.group_code                = g.group_code
          AND g.genusspecies_name         = b.genus_species
          AND g.gpc_application           = b.gpc_application
          AND b.match_on_x                = 'GS'
          AND p.brick_code IS NULL;
    """,

    "Q3_G_null": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = b.brick_code
        FROM "_import_Bricks_With_Genus_Species" AS b,
             "_import_Groups_With_Genus_Species" AS g,
             "Plant_Genus_Species"       AS s
        WHERE p.plant_registration_number = s.plant_registration_number
          AND p.group_code                = g.group_code
          AND b.gpc_application           = g.gpc_application
          AND b.genus_species             = s.latin_genus_name
          AND b.match_on_x                = 'G'
          AND p.brick_code IS NULL;
    """,

    "Q4_G_null_species": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = b.brick_code
        FROM "Plant_Genus_Species"       AS s,
             "_import_Groups_With_Genus_Species" AS g,
             "_import_Bricks_With_Genus_Species" AS b
        WHERE p.plant_registration_number = s.plant_registration_number
          AND p.group_code                = g.group_code
          AND b.gpc_application           = g.gpc_application
          AND b.genus_species             = g.genus_name
          AND b.match_on_x                = 'G'
          AND p.brick_code IS NULL;
    """,

    "Q5_name_GS": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = b.brick_code
        FROM
            "Product_genus_species_from_Product_name"   AS pg,
            "_import_Groups_With_Genus_Species"         AS g,
            "_import_Bricks_With_Genus_Species"         AS b
        WHERE
            pg.product_id             = p.product_id           
        AND g.group_code            = p.group_code           
        AND b.genus_species         = pg.genus_species       
        AND b.gpc_application       = g.gpc_application      
        AND b.match_on_x            = 'GS'                   
        AND p.brick_code IS NULL;  
    """,
    "Q6_name_G": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = b.brick_code
        FROM "Product_genus_species_from_Product_name" AS pg
        JOIN "_import_Groups_With_Genus_Species" AS g
        ON pg."gpc_application" = g."gpc_application"
        JOIN "_import_Bricks_With_Genus_Species" AS b
        ON pg."genus" = b.genus_species
        WHERE
            p.product_id           = pg.product_id    
        AND p.group_code         = g.group_code    
        AND p.brick_code IS NULL                    
        AND b.match_on_x        = 'G';              
    """,
    "Q7_group_default": """
        UPDATE "PRODUCT_GPC" AS p
        SET brick_code = g.brick_code_through_group_code
        FROM "_import_Groups_With_Genus_Species" g
        WHERE p.group_code = g.group_code
          AND p.brick_code IS NULL
          AND g.brick_code_through_group_code IS NOT NULL;
    """,
    "Q8_default_values": """
        UPDATE "PRODUCT_GPC"
        SET brick_code = CASE
            WHEN group_code < 10700000 THEN 10006502
            WHEN group_code > 20000000 THEN 10006679
            ELSE 10006547
        END
        WHERE brick_code IS NULL;
    """
}

def parse_filename_date(fname: str) -> str:
    code = fname[-10:-4]
    try:
        return datetime.strptime(code, "%d%m%y").strftime("%d-%m-%Y")
    except:
        return ""

def clear_old_data(ws):
    for row in ws.iter_rows(min_row=START_ROW, min_col=START_COL):
        for cell in row:
            cell.value = None

def fill_florecompc():
    wb = load_workbook(TEMPLATE_PATH)
    for sheet_name, pattern in SHEETS_WITH_CODELIST.items():
        ws = wb[sheet_name]
        clear_old_data(ws)
        files = sorted(DATA_DIR.glob(pattern))
        if not files:
            print(f"⚠️ Geen bestanden voor '{sheet_name}'")
            continue
        txt = files[-1]
        print(f"📥 Vul '{sheet_name}' met '{txt.name}'")
        try:
            df = pd.read_csv(txt, sep=';', encoding='utf-8', header=None, dtype=str)
        except:
            df = pd.read_csv(txt, sep=';', encoding='latin-1', header=None, dtype=str)
        for r, row in enumerate(df.itertuples(index=False, name=None), START_ROW):
            for c, v in enumerate(row, START_COL):
                ws.cell(row=r, column=c, value=v)
        ws.cell(row=1, column=5, value=txt.name)
        ws.cell(row=2, column=5, value=parse_filename_date(txt.name))
    wb.save(FILLED_PATH)
    print(f"✅ Gevuld: {FILLED_PATH}")
def split_genus_species(s):
    """
    Given a combined_product string `s`, return a Series with:
      - genus_species: either "Genus Species" or just "Genus" if the second token ends with 'Grp'
      - genus: the first token
    """
    # handle empty / NaN
    if not s or pd.isna(s) or s.strip() == "":
        return pd.Series({"genus_species": None, "genus": None})

    # split on whitespace, max 2 parts
    parts = s.strip().split(None, 2)
    genus = parts[0]

    # decide if we include the second token
    if len(parts) > 1:
        second = parts[1]
        # if the second token ends with 'Grp' (a group name), drop it
        if second.lower().endswith("grp"):
            genus_species = genus
        else:
            genus_species = f"{genus} {second}"
    else:
        genus_species = genus

    return pd.Series({
        "genus_species": genus_species,
        "genus":         genus
    })

def load_to_postgres():
    from sqlalchemy import create_engine, text

    engine = create_engine(DB_CONNECTION_STRING)

    # 0) Drop view so we can safely replace its base tables
    print("🔧 Droppen view Plant_Genus_Species (indien aanwezig)")
    with engine.begin() as conn:
        conn.execute(text('DROP VIEW IF EXISTS "Plant_Genus_Species"'))

    # 1a) _import_PLANT
    print("📥 Laden sheet 'Gewas' → tabel '_import_PLANT'")
    df_plant = (
        pd.read_excel(FILLED_PATH, sheet_name="Gewas", header=2, dtype=str)
          .fillna("")
    )
    df_plant = normalize_and_translate(df_plant)
    df_plant.to_sql("_import_PLANT", engine, if_exists="replace", index=False)
    print(f"✅ _import_PLANT geladen ({len(df_plant)} rijen)")

    # 1b) _import_GENUS
    print("📥 Laden sheet 'Geslacht' → tabel '_import_GENUS'")
    df_genus = (
        pd.read_excel(FILLED_PATH, sheet_name="Geslacht", header=2, dtype=str)
          .fillna("")
    )
    df_genus = normalize_and_translate(df_genus)
    df_genus = df_genus.rename(columns={"naam": "latin_genus_name"})
    df_genus.to_sql("_import_GENUS", engine, if_exists="replace", index=False)
    print(f"✅ _import_GENUS geladen ({len(df_genus)} rijen)")

    # 1c) _import_SPECIES
    print("📥 Laden sheet 'Soort' → tabel '_import_SPECIES'")
    df_species = (
        pd.read_excel(FILLED_PATH, sheet_name="Soort", header=2, dtype=str)
          .fillna("")
    )
    df_species = normalize_and_translate(df_species)
    df_species = df_species.rename(columns={"latin_genus_name": "latin_species_name"})
    df_species.to_sql("_import_SPECIES", engine, if_exists="replace", index=False)
    print(f"✅ _import_SPECIES geladen ({len(df_species)} rijen)")

    # 1d) PRODUCT_GPC
    print("📥 Laden sheet 'Product' → tabel 'PRODUCT_GPC'")
    df_prod = (
        pd.read_excel(FILLED_PATH, sheet_name="Product", header=2, dtype=str)
          .fillna("")
    )
    df_prod = normalize_and_translate(df_prod)
    df_prod = df_prod.rename(columns={
        "groepscode": "group_code",
        "productid": "product_id"
    })
    # — Cast IDs to integers so Postgres sees BIGINT on both sides of the JOIN —
    df_prod["group_code"] = pd.to_numeric(df_prod["group_code"], downcast="integer")
    df_prod["product_id"] = pd.to_numeric(df_prod["product_id"], downcast="integer")
    df_prod.to_sql("PRODUCT_GPC", engine, if_exists="replace", index=False)
    print(f"✅ PRODUCT_GPC geladen ({len(df_prod)} rijen)")
    # ──────────────────────────────────────────────────────────────────────────────
    # 1e) Build Product_genus_species_from_Product_name (w/ gpc_application)
    
    print("🔧 Aanmaken tabel Product_genus_species_from_Product_name")

    # (1) first, build the simple genus<->species from the combined_product text
    pg = df_prod[["product_id", "combined_product"]].copy()
    pg[["genus_species", "genus"]] = pg["combined_product"].apply(split_genus_species)
    pg = pg[["product_id", "genus_species", "genus"]]

    # (2) now read in your Groups export so we can grab the gpc_application
    raw_groups = pd.read_csv(
        DATA_DIR / "_import_Groups_With_Genus_Species.txt",
        sep=";", header=0, dtype=str
    ).fillna("")

    # sheet‐specific fix: rename the hyphened and uppercase columns
    raw_groups = raw_groups.rename(columns={
        "genus-species_name": "genusspecies_name",
        "GPC_application":    "gpc_application"
    })

    # normalize & translate
    df_groups = normalize_and_translate(raw_groups)

    # cast group_code (if you’ll ever need it) and ensure gpc_application is text
    df_groups["group_code"] = pd.to_numeric(df_groups["group_code"], errors="coerce")
    df_groups["gpc_application"] = pd.to_numeric(
        df_groups["gpc_application"],
        errors="coerce"
    ).astype("Int64")

    # (3) merge it into pg on the genus_species key
    pg = pg.merge(
        df_groups[["genusspecies_name", "gpc_application"]],
        left_on="genus_species", right_on="genusspecies_name",
        how="left"
    )
    pg["gpc_application"] = pg["gpc_application"].astype("Int64")
    # drop the extra column now that we have gpc_application
    pg = pg[["product_id", "genus_species", "genus", "gpc_application"]]

    # (4) write out to Postgres
    pg.to_sql(
        "Product_genus_species_from_Product_name",
        engine, if_exists="replace", index=False
    )
    print(f"✅ Product_genus_species_from_Product_name aangemaakt ({len(pg)} rijen)")
 # 1e) Bricks & Groups exports
    for name in ("Bricks_With_Genus_Species", "Groups_With_Genus_Species"):
        txt = DATA_DIR / f"_import_{name}.txt"
        print(f"📥 Laden export '{name}' → tabel '{name}'")

        raw = pd.read_csv(txt, sep=";", header=0, dtype=str).fillna("")
        if "GPC_application" in raw.columns:
            raw = raw.rename(columns={"GPC_application": "gpc_application"})
            print("   ▶ Kolom 'GPC_application' hernoemd naar 'gpc_application'")
        if "GPC_application" in raw.columns:
            raw = raw.rename(columns={"Match_on_x": "match_on_x"})
            print("   ▶ Kolom 'Match_on_x' hernoemd naar 'match_on_x'")
        # rename the hyphened header before normalize
        if name == "Groups_With_Genus_Species" and "genus-species_name" in raw.columns:
            raw = raw.rename(columns={"genus-species_name": "genusspecies_name"})
            print("   ▶ Kolom genusspecies_name voorbereid in DataFrame")

        df = normalize_and_translate(raw)

        # cast numeric keys…
        if name == "Groups_With_Genus_Species":
                # cast gpc_application to Int64 before to_sql
 
            df["gpc_application"] = pd.to_numeric(
                df["gpc_application"], errors="coerce"
            ).astype("Int64")
            df["group_code"]      = pd.to_numeric(
                df["group_code"], errors="coerce"
            ).astype("Int64")
            df["brick_code_through_group_code"] = pd.to_numeric(
                df["brick_code_through_group_code"], errors="coerce"
            )

        df.to_sql(name, engine, if_exists="replace", index=False)
        print(f"✅ {name} geladen ({len(df)} rijen)")
        print(df.columns)

        # **force a DB‐level rename** in case the old column survived


    with engine.begin() as conn:
        print("🔧 Aanmaken view Plant_Genus_Species")
        conn.execute(text(VIEW_DDLS["view_plant_genus_species"]))

        print("🔧 Toevoegen kolom PRODUCT_GPC.brick_code")
        conn.execute(text(
            'ALTER TABLE "PRODUCT_GPC" ADD COLUMN IF NOT EXISTS brick_code bigint'
        ))

        print("▶ Uitvoeren update-queries")
        for uname, sql in UPDATE_QUERIES.items():
            print(f"   🔄 {uname}")
            conn.execute(text(sql))

    print("✅ Database en alle queries zijn succesvol uitgevoerd.")

CODE_LIST_NUMBERS = {
    "SEGMENT":                             30,
    "FAMILY":                              31,
    "CLASS":                               32,
    "BRICK":                               33,
    "ATTRIBUTE_TYPE":                      34,
    "ATTRIBUTE_VALUE":                     35,
    "SEGMENT_FAMILY_CLASS_BRICK":          36,
    "BRICK_ATTRIBUTE_TYPE_ATTRIBUTE_VALUE":37,
    "NAME_GPC":                            21,
    "NAME_TYPE":                           22,
    "PRODUCT_GPC":                         39,
    "COLOR":                               20,
    "APPLICATION":                         15,
    # add more if needed…
}

def export_code_lists():
    engine = create_engine(DB_CONNECTION_STRING)
    insp = inspect(engine)
    today = datetime.today().strftime("%Y%m%d")

    for table, code_num in CODE_LIST_NUMBERS.items():
        if table not in insp.get_table_names(schema="public"):
            print(f"⚠️  skipping {table!r}: not found in database")
            continue

        print(f"📤 Exporting table {table} → C{code_num}_{today}.txt")

        # read entire table
        df = pd.read_sql_table(table, engine, schema="public")

        # drop the GS1 code_list_id column itself if present
        if "code_list_id" in df.columns:
            df = df.drop(columns=["code_list_id"])

        # output file
        out_path = OUTPUT_DIR / f"C{code_num}_{today}.txt"

        # write out with ; separator, no header, no index, no quoting
        df.to_csv(
            out_path,
            sep=";",
            index=False,
            header=False,
            quoting=csv.QUOTE_NONE,
            escapechar="\\",
        )

        print(f"   ✅ Wrote {len(df)} rows to {out_path.name}")
if __name__ == "__main__":
    #load_to_postgres()
    
    export_code_lists()