
import math
import datetime as dt
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# --- 1. Bestands­locatie -----------------------------------------------------
# Pas dit pad aan naar jouw eigen bestand:
FILE_PATH = r"C:\Users\c.elkhattabi\Desktop\Rapport 2025-07-17 10-02-21.xlsx"
SHEET_NAME = "Rapport"         # of een andere werkbladnaam
# ── 1. Werkblad openen (read-only, alleen berekende waarden) ────────────────


def to_hours(x):
    if x is None or x == "" or (isinstance(x, float) and math.isnan(x)):
        return math.nan
    if isinstance(x, (dt.timedelta, pd.Timedelta)):
        return x.total_seconds()/3600
    if isinstance(x, dt.time):
        return x.hour + x.minute/60 + x.second/3600
    if isinstance(x, dt.datetime):
        dagen = x.day - 1
        return dagen*24 + x.hour + x.minute/60 + x.second/3600
    if isinstance(x, str):
        try:
            return pd.to_timedelta(x).total_seconds()/3600
        except Exception:
            pass
    return math.nan

def load_hours() -> pd.DataFrame:
    wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb.active                                 # of wb["Rapport"] als je wilt

    # ── 2. Header-rij vinden ----------------------------------------------------
    header_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "Taak":               # kolom A == "Taak"
            header = list(row)
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("'Taak' niet gevonden – check je werkblad")

    # ── 3. Data onder de header binnenhalen tot eerste lege regel  --------------
    records = []
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        if not any(row):                           # helemaal leeg? → klaar
            break
        records.append(row)

    # ── 4. DataFrame bouwen -----------------------------------------------------
    df = pd.DataFrame(records, columns=header)
    df["Uren"] = df["Duur"].apply(to_hours)
    print(df)
    return df
load_hours()