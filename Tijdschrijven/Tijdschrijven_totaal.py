#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
lees_tijdschrijven_rapport.py
─────────────────────────────
• zoekt de laagste (laatste) kopregel 'Taak'
• vult samengevoegde lege cellen (Taak / Klant / Project / Afdeling) met ffill
• rekent kolom 'Duur' om naar decimale uren
• toont snelle totalen
"""

import math
import datetime as dt
from pathlib import Path
from Tijdschrijven.file import main as file
import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------- SETTINGS
FILE_PATH  = file
SHEET_NAME = "Rapport"
# -------------------------------------------------------------------------


def to_hours(x) -> float:
    """ Zet tijd / duur → float-uren. """
    if x is None or x == "" or (isinstance(x, float) and math.isnan(x)):
        return math.nan
    if isinstance(x, (dt.timedelta, pd.Timedelta)):
        return x.total_seconds() / 3600
    if isinstance(x, dt.time):
        return x.hour + x.minute / 60 + x.second / 3600
    if isinstance(x, dt.datetime):  # Excel-waarde > 24u
        days = x.day - 1
        return days * 24 + x.hour + x.minute / 60 + x.second / 3600
    if isinstance(x, str):
        try:
            return pd.to_timedelta(x).total_seconds() / 3600
        except Exception:
            pass
    return math.nan


def load_hours(path: Path = FILE_PATH, sheet: str = SHEET_NAME) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active

    # ── 1. zoek de LAATSTE rij waar kolom A 'Taak' is ────────────────────
    header_row, header = None, None
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "Taak":
            header_row, header = ridx, list(row)  # overschrijft telkens
    if header_row is None:
        raise ValueError("'Taak' niet gevonden in werkblad")

    # ── 2. lees alle rijen onder de header tot eerste volledig lege rij ──
    records = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(r):
            break
        records.append(r)

    # ── 3. DataFrame + ffill op samengevoegde kolommen ───────────────────
    df = pd.DataFrame(records, columns=header)
    cols_ffill = ["Taak", "Klant", "Project", "Afdeling"]
    df[cols_ffill] = df[cols_ffill].ffill()

    # ── 4. uren toevoegen ────────────────────────────────────────────────
    df["Uren"] = df["Duur"].apply(to_hours)
    return df

def main_tijd():
    df = load_hours()

    # Top-taken
    uren_per_taak = (
        df.groupby("Taak", as_index=False)["Uren"]
        .sum()
        .sort_values("Uren", ascending=False)
    )
    return uren_per_taak


